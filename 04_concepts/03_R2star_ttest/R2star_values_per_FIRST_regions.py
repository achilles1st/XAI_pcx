import os

import json
import nibabel as nib
import numpy as np
from openpyxl import Workbook

from scipy import stats

this_file_path = os.path.dirname(os.path.realpath(__file__))
data_base_path = os.path.join(this_file_path, '../../02_models_training/00_data')

with open('../../01_data/04_matching/ASPSF_matched.json', 'r') as infile:
  ASPSF_scans = json.load(infile)

with open('../../01_data/04_matching/ProDem_matched.json', 'r') as infile:
  ProDem_scans = json.load(infile)

subject_dirs = ASPSF_scans + ProDem_scans
subject_dirs.sort()

wb = Workbook()
wb.remove(wb.active)

for config in [
  # image             masks
  ('R2star.nii.gz',   'T1_1mm_first_masks_@_R2star_dof6')
]:
  (image_name, masks_path) = config

  ws = wb.create_sheet(image_name.split('.')[0])
  ws.freeze_panes = ws['B2']
  ws.cell(1, 1, value='Subject')
  
  ASPSF_data = []
  ProDem_data = []

  subject_number = 0
  for subject_dir in subject_dirs:
    if subject_dir in [
      '801129__20130307',
    ]:
      continue

    subject_data = []
    if subject_dir.startswith('8'):
      ProDem_data.append(subject_data)
    else:
      ASPSF_data.append(subject_data)

    image_path = os.path.join(data_base_path, subject_dir, image_name)
    image = nib.load(image_path)
    image_data = image.get_fdata()[:, :, :]

    mask_number = 0
    scalars_per_mask = 6
    masks_list = [
      (['-L_Caud_first.nii.gz', '-L_Pall_first.nii.gz', '-L_Puta_first.nii.gz'], ['-R_Caud_first.nii.gz', '-R_Pall_first.nii.gz', '-R_Puta_first.nii.gz']),
      (['-L_Pall_first.nii.gz', '-L_Puta_first.nii.gz', '-L_Thal_first.nii.gz'], ['-R_Pall_first.nii.gz', '-R_Puta_first.nii.gz', '-R_Thal_first.nii.gz']),
      (['-L_Accu_first.nii.gz'], ['-R_Accu_first.nii.gz']),
      (['-L_Amyg_first.nii.gz'], ['-R_Amyg_first.nii.gz']),
      (['-L_Caud_first.nii.gz'], ['-R_Caud_first.nii.gz']),
      (['-L_Hipp_first.nii.gz'], ['-R_Hipp_first.nii.gz']),
      (['-L_Pall_first.nii.gz'], ['-R_Pall_first.nii.gz']),
      (['-L_Puta_first.nii.gz'], ['-R_Puta_first.nii.gz']),
      (['-L_Thal_first.nii.gz'], ['-R_Thal_first.nii.gz']),
    ]
    for (left_mask_names, right_mask_names) in masks_list:
      # left
      left_column_names = []
      left_region_datas = []
      for left_mask_name in left_mask_names:
        left_mask_path = os.path.join(data_base_path, subject_dir, masks_path, left_mask_name)
        left_mask = nib.load(left_mask_path)
        if np.any(image.affine != left_mask.affine):
          print(subject_dir + ' affines differ!')

        left_mask_data = left_mask.get_fdata()
        left_region_data_where = left_mask_data > 0.
        
        left_column_names.append(left_mask_name[0:7])
        left_region_datas.append(image_data[left_region_data_where])

      # right
      right_column_names = []
      right_region_datas = []
      for right_mask_name in right_mask_names:
        right_mask_path = os.path.join(data_base_path, subject_dir, masks_path, right_mask_name)
        right_mask = nib.load(right_mask_path)
        if np.any(image.affine != right_mask.affine):
          print(subject_dir + ' affines differ!')

        right_mask_data = right_mask.get_fdata()
        right_region_data_where = right_mask_data > 0.
        
        right_column_names.append(right_mask_name[0:7])
        right_region_datas.append(image_data[right_region_data_where])

      # header
      ws.cell(1, mask_number * scalars_per_mask + 2, value=','.join(left_column_names) + ' Median')
      ws.cell(1, mask_number * scalars_per_mask + 3, value=','.join(left_column_names) + ' Mean')
      ws.cell(1, mask_number * scalars_per_mask + 4, value=','.join(right_column_names) + ' Median')
      ws.cell(1, mask_number * scalars_per_mask + 5, value=','.join(right_column_names) + ' Mean')
      ws.cell(1, mask_number * scalars_per_mask + 6, value=','.join([right_column_name[3:7] for right_column_name in right_column_names]) + ' Median')
      ws.cell(1, mask_number * scalars_per_mask + 7, value=','.join([right_column_name[3:7] for right_column_name in right_column_names]) + ' Mean')

      ws.cell(subject_number + 2, 1, value=subject_dir)
      
      left_median = np.mean([np.median(left_region_data) for left_region_data in left_region_datas])
      left_mean = np.mean([np.mean(left_region_data) for left_region_data in left_region_datas])
      subject_data.append(left_median)
      subject_data.append(left_mean)

      ws.cell(subject_number + 2, mask_number * scalars_per_mask + 2, value=left_median)
      ws.cell(subject_number + 2, mask_number * scalars_per_mask + 3, value=left_mean)

      right_median = np.mean([np.median(right_region_data) for right_region_data in right_region_datas])
      right_mean = np.mean([np.mean(right_region_data) for right_region_data in right_region_datas])
      subject_data.append(right_median)
      subject_data.append(right_mean)

      ws.cell(subject_number + 2, mask_number * scalars_per_mask + 4, value=right_median)
      ws.cell(subject_number + 2, mask_number * scalars_per_mask + 5, value=right_mean)

      median = np.mean([left_median, right_median])
      mean = np.mean([left_mean, right_mean])
      subject_data.append(median)
      subject_data.append(mean)

      ws.cell(subject_number + 2, mask_number * scalars_per_mask + 6, value=median)
      ws.cell(subject_number + 2, mask_number * scalars_per_mask + 7, value=mean)
      
      mask_number = mask_number + 1

    subject_number = subject_number + 1
  
  (ttest_t_statistic, ttest_p_value) = stats.ttest_ind(ASPSF_data, ProDem_data, axis=0, equal_var=False)
  (mwu_t_statistic, mwu_p_value) = stats.mannwhitneyu(ASPSF_data, ProDem_data, axis=0)
  ASPSF_median = np.median(ASPSF_data, axis=0)
  ProDem_median = np.median(ProDem_data, axis=0)
  (ASPSF_1q, ASPSF_3q) = np.percentile(ASPSF_data, (25, 75), axis=0)
  (ProDem_1q, ProDem_3q) = np.percentile(ProDem_data, (25, 75), axis=0)
  ASPSF_skew = stats.skew(ASPSF_data, axis=0)
  ASPSF_kurtosis = stats.kurtosis(ASPSF_data, axis=0)
  ProDem_skew = stats.skew(ProDem_data, axis=0)
  ProDem_kurtosis = stats.kurtosis(ProDem_data, axis=0)

  ASPSF_ks_p = []
  ProDem_ks_p = []
  for i in range(scalars_per_mask * len(masks_list)):
    # result = stats.kstest(np.asarray(ASPSF_data)[:, i], 'norm')
    (t, p) = stats.shapiro(np.asarray(ASPSF_data)[:, i])
    ASPSF_ks_p.append(p)
    # (t, p) = stats.kstest(np.asarray(ProDem_data)[:, i], 'norm')
    (t, p) = stats.shapiro(np.asarray(ProDem_data)[:, i])
    ProDem_ks_p.append(p)

  ws.cell(subject_number + 4, 1, value='ttest t statistic')
  ws.cell(subject_number + 5, 1, value='ttest p value')
  ws.cell(subject_number + 6, 1, value='mwu t statistic')
  ws.cell(subject_number + 7, 1, value='mwu p value')
  ws.cell(subject_number + 8, 1, value='ASPSF median')
  ws.cell(subject_number + 9, 1, value='ProDem median')
  ws.cell(subject_number + 10, 1, value='ASPSF Q1')
  ws.cell(subject_number + 11, 1, value='ProDem Q1')
  ws.cell(subject_number + 12, 1, value='ASPSF Q3')
  ws.cell(subject_number + 13, 1, value='ProDem Q3')
  ws.cell(subject_number + 14, 1, value='ASPSF skew')
  ws.cell(subject_number + 15, 1, value='ASPSF kurtosis')
  ws.cell(subject_number + 16, 1, value='ProDem skew')
  ws.cell(subject_number + 17, 1, value='ProDem kurtosis')
  ws.cell(subject_number + 18, 1, value='ASPSF shapiro-wilk p value')
  ws.cell(subject_number + 19, 1, value='ProDem shapiro-wilk p value')
  for i in range(len(ttest_t_statistic)):
    ws.cell(subject_number + 4, i + 2, value=ttest_t_statistic[i])
    ws.cell(subject_number + 5, i + 2, value=ttest_p_value[i])
    ws.cell(subject_number + 6, i + 2, value=mwu_t_statistic[i])
    ws.cell(subject_number + 7, i + 2, value=mwu_p_value[i])
    ws.cell(subject_number + 8, i + 2, value=ASPSF_median[i])
    ws.cell(subject_number + 9, i + 2, value=ProDem_median[i])
    ws.cell(subject_number + 10, i + 2, value=ASPSF_1q[i])
    ws.cell(subject_number + 11, i + 2, value=ProDem_1q[i])
    ws.cell(subject_number + 12, i + 2, value=ASPSF_3q[i])
    ws.cell(subject_number + 13, i + 2, value=ProDem_3q[i])
    ws.cell(subject_number + 14, i + 2, value=ASPSF_skew[i])
    ws.cell(subject_number + 15, i + 2, value=ASPSF_kurtosis[i])
    ws.cell(subject_number + 16, i + 2, value=ProDem_skew[i])
    ws.cell(subject_number + 17, i + 2, value=ProDem_kurtosis[i])
    ws.cell(subject_number + 18, i + 2, value=ASPSF_ks_p[i])
    ws.cell(subject_number + 19, i + 2, value=ProDem_ks_p[i])

    ws.cell(subject_number + 22, i + 2, value='{0:.2f} ({1:.2f}-{2:.2f})'.format(ASPSF_median[i], ASPSF_1q[i], ASPSF_3q[i]))
    ws.cell(subject_number + 23, i + 2, value='{0:.2f} ({1:.2f}-{2:.2f})'.format(ProDem_median[i], ProDem_1q[i], ProDem_3q[i]))
    ws.cell(subject_number + 24, i + 2, value='{0:.3f}'.format(ttest_p_value[i]))
    ws.cell(subject_number + 25, i + 2, value='{0:.3f}'.format(mwu_p_value[i]))

wb.save('R2star_values_per_FIRST_regions.xlsx')